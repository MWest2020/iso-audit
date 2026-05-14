"""Tabulaire export van audit-bevindingen — CSV + Excel (`.xlsx`).

Naast het Markdown-rapport (`local_report.py`) levert deze module een platte
tabelvorm geschikt voor filtering, Sheets, en audit-trail. Thema-bundeling
hergebruikt de heuristiek uit `iso_audit.classification.thema` (route A);
optioneel kan een caller `llm_themas`-mapping meegeven voor route B.

Gemigreerd uit `Ops_to_Biz/audit/tabular_report.py` per milestone B §2.5.1.
Wijzigingen: `THEMA_LIJST`/`THEMA_REGELS`/`bepaal_thema` zijn niet meer
gedupliceerd — alle theme-resolutie via
`iso_audit.classification.thema.bepaal_thema`. `audit.normteksten` →
`iso_audit.data.normteksten`. Path-resolution via `Path`.
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from iso_audit.classification.thema import bepaal_thema
from iso_audit.data.normteksten import NORMTEKSTEN_9001, NORMTEKSTEN_27001

logger = logging.getLogger(__name__)

# src/iso_audit/reporting/tabular_report.py → parent x 4 = repo root.
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
_BASE_OUTPUT_DIR = _REPO_ROOT / "output" / "audit_reports"
DEFAULT_OUTPUT_DIR = str(_BASE_OUTPUT_DIR)
DEFAULT_DB_PATH = str(_REPO_ROOT / "output" / "audit.db")

VOLGORDE_CLASSIFICATIE: dict[str, int] = {"NC": 0, "OFI": 1, "positief": 2}


def _audit_run_dir() -> str:
    """Run-dir-conventie; identiek aan `local_report._audit_run_dir`."""
    explicit = os.environ.get("LOCAL_REPORT_DIR")
    if explicit:
        return explicit
    run_id = os.environ.get("AUDIT_RUN_ID") or f"audit_{date.today().isoformat()}"
    return str(_BASE_OUTPUT_DIR / run_id)


def _bepaal_norm_voor_clausule(clausule_id: str) -> str:
    """`"9001"` / `"27001"` / `"beide"` / `"onbekend"` op basis van normteksten."""
    in_9001 = clausule_id in NORMTEKSTEN_9001
    in_27001 = clausule_id in NORMTEKSTEN_27001
    if in_9001 and in_27001:
        return "beide"
    if in_9001:
        return "9001"
    if in_27001:
        return "27001"
    return "onbekend"


def _doc_url(bevinding: dict[str, Any]) -> str:
    """Drive of Miro-URL afgeleid uit `doc_id` + `herkomst`."""
    doc_id = bevinding.get("doc_id", "")
    if not doc_id:
        return ""
    if bevinding.get("herkomst") == "Miro":
        board_id = os.environ.get("MIRO_BOARD_ID", "")
        if board_id:
            return f"https://miro.com/app/board/{board_id}/?moveToWidget={doc_id}"
        return ""
    return f"https://drive.google.com/file/d/{doc_id}/view"


def _verrijk(
    bevindingen: list[dict[str, Any]],
    llm_themas: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Voeg afgeleide kolommen toe: `norm`, `thema`, `thema_bron`, `doc_url`.

    `llm_themas` is een optionele mapping van `bev_id` → thema. Wanneer een
    bevinding daarin staat, wordt het LLM-thema gebruikt en is `thema_bron`
    gelijk aan `"llm"`. Anders valt route A (heuristiek) in.
    """
    verrijkt: list[dict[str, Any]] = []
    for i, bev in enumerate(bevindingen):
        clausule = bev.get("clausule") or bev.get("clausule_id", "")
        bev_id = str(bev.get("id") or bev.get("_bev_id") or i)
        row = dict(bev)
        row["_bev_id"] = bev_id
        row["clausule"] = clausule
        row["norm"] = _bepaal_norm_voor_clausule(clausule)
        if llm_themas and bev_id in llm_themas:
            row["thema"] = llm_themas[bev_id]
            row["thema_bron"] = "llm"
        else:
            row["thema"] = bepaal_thema(bev)
            row["thema_bron"] = "heuristisch"
        row["doc_url"] = _doc_url(bev)
        verrijkt.append(row)
    return verrijkt


def _clausule_sleutel(c: str) -> tuple[int, ...]:
    """Numerieke sortering: `"10.2"` → `(10, 2)`; ongeldig → `(9999,)`."""
    delen = re.findall(r"\d+", c)
    return tuple(int(d) for d in delen) if delen else (9999,)


def _sorteer(bevindingen: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sorteer: clausule numeriek → classificatie (NC/OFI/positief) → thema."""
    return sorted(
        bevindingen,
        key=lambda b: (
            _clausule_sleutel(b.get("clausule", "")),
            VOLGORDE_CLASSIFICATIE.get(b.get("classificatie", ""), 9),
            b.get("thema", ""),
        ),
    )


def _bestandsnaam(prefix: str, norm: str, scherpte: float, extensie: str) -> str:
    norm_deel = norm.replace(" ", "").replace("+", "-")
    scherpte_label = f"_s{str(scherpte).replace('.', '')}" if scherpte != 1.0 else ""
    return f"{prefix}_{norm_deel}_{date.today()}{scherpte_label}.{extensie}"


CSV_KOLOMMEN: list[str] = [
    "norm",
    "clausule",
    "clausule_titel",
    "classificatie",
    "thema",
    "thema_bron",
    "document_naam",
    "herkomst",
    "beschrijving",
    "onderbouwing",
    "doc_id",
    "doc_url",
    "classified_at",
]


def schrijf_csv(
    bevindingen: list[dict[str, Any]],
    norm: str,
    scherpte: float = 1.0,
    output_dir: str | None = None,
    llm_themas: dict[str, str] | None = None,
) -> str:
    """Schrijf bevindingen als platte CSV. Returnt het pad."""
    out_dir = Path(output_dir) if output_dir else Path(_audit_run_dir())
    out_dir.mkdir(parents=True, exist_ok=True)
    pad = out_dir / _bestandsnaam("Bevindingen", norm, scherpte, "csv")
    rijen = _sorteer(_verrijk(bevindingen, llm_themas=llm_themas))

    with pad.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_KOLOMMEN, extrasaction="ignore")
        writer.writeheader()
        for r in rijen:
            writer.writerow({k: r.get(k, "") for k in CSV_KOLOMMEN})
    logger.info("CSV geschreven: %s (%d rijen)", pad, len(rijen))
    return str(pad)


CLASSIFICATIE_KLEUREN: dict[str, str] = {
    "NC": "FFC7CE",  # rood
    "OFI": "FFEB9C",  # oranje
    "positief": "C6EFCE",  # groen
}


def schrijf_excel(
    bevindingen: list[dict[str, Any]],
    norm: str,
    scherpte: float = 1.0,
    output_dir: str | None = None,
    llm_themas: dict[str, str] | None = None,
) -> str:
    """Schrijf bevindingen als `.xlsx` met drie tabs.

    1. **Samenvatting** — aantal per (norm, thema, classificatie).
    2. **Bevindingen** — platte tabel + conditional-formatting per NC/OFI/pos.
    3. **Per clausule** — aantal NC/OFI/positief per clausule.
    """
    out_dir = Path(output_dir) if output_dir else Path(_audit_run_dir())
    out_dir.mkdir(parents=True, exist_ok=True)
    pad = out_dir / _bestandsnaam("Bevindingen", norm, scherpte, "xlsx")
    rijen = _sorteer(_verrijk(bevindingen, llm_themas=llm_themas))

    wb = Workbook()
    _tab_samenvatting(wb.active, rijen)
    _tab_bevindingen(wb.create_sheet("Bevindingen"), rijen)
    _tab_per_clausule(wb.create_sheet("Per clausule"), rijen)
    wb.save(str(pad))
    logger.info("Excel geschreven: %s (%d rijen)", pad, len(rijen))
    return str(pad)


def _tab_samenvatting(ws: Any, rijen: list[dict[str, Any]]) -> None:
    ws.title = "Samenvatting"
    header_font = Font(bold=True)

    ws.append(["Totalen"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append(["Classificatie", "Aantal"])
    for c in ws[2]:
        c.font = header_font
    teller_cls: Counter[str] = Counter(r["classificatie"] for r in rijen)
    for cls in ("NC", "OFI", "positief"):
        ws.append([cls, teller_cls.get(cls, 0)])
    ws.append(["Totaal", len(rijen)])
    ws[ws.max_row][0].font = header_font
    ws[ws.max_row][1].font = header_font
    ws.append([])

    ws.append(["Bundeling per thema"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Thema", "NC", "OFI", "Positief", "Totaal"])
    for c in ws[ws.max_row]:
        c.font = header_font
    teller_thema: dict[str, Counter[str]] = defaultdict(Counter)
    for r in rijen:
        teller_thema[r["thema"]][r["classificatie"]] += 1

    def thema_sleutel(item: tuple[str, Counter[str]]) -> tuple[bool, int, str]:
        thema, teller = item
        totaal = sum(teller.values())
        return (thema == "Overig", -totaal, thema)

    for thema, teller in sorted(teller_thema.items(), key=thema_sleutel):
        ws.append(
            [
                thema,
                teller.get("NC", 0),
                teller.get("OFI", 0),
                teller.get("positief", 0),
                sum(teller.values()),
            ]
        )
    ws.append([])

    ws.append(["Per norm"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["Norm", "NC", "OFI", "Positief", "Totaal"])
    for c in ws[ws.max_row]:
        c.font = header_font
    teller_norm: dict[str, Counter[str]] = defaultdict(Counter)
    for r in rijen:
        teller_norm[r["norm"]][r["classificatie"]] += 1
    for norm_key in ("9001", "27001", "beide", "onbekend"):
        if norm_key in teller_norm:
            t = teller_norm[norm_key]
            ws.append(
                [
                    norm_key,
                    t.get("NC", 0),
                    t.get("OFI", 0),
                    t.get("positief", 0),
                    sum(t.values()),
                ]
            )
    _auto_width(ws)


def _tab_bevindingen(ws: Any, rijen: list[dict[str, Any]]) -> None:
    ws.append(CSV_KOLOMMEN)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")

    for r in rijen:
        ws.append([r.get(k, "") for k in CSV_KOLOMMEN])

    cls_col_idx = CSV_KOLOMMEN.index("classificatie") + 1
    cls_letter = get_column_letter(cls_col_idx)
    rng = f"{cls_letter}2:{cls_letter}{ws.max_row}"
    for cls, kleur in CLASSIFICATIE_KLEUREN.items():
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=[f'"{cls}"'],
                fill=PatternFill("solid", fgColor=kleur),
            ),
        )
    for kolom_naam in ("beschrijving", "onderbouwing"):
        idx = CSV_KOLOMMEN.index(kolom_naam) + 1
        for row_cells in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for c in row_cells:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "C2"
    _auto_width(ws, max_breedte=60)


def _tab_per_clausule(ws: Any, rijen: list[dict[str, Any]]) -> None:
    ws.append(["Clausule", "Norm", "Titel", "NC", "OFI", "Positief", "Totaal"])
    for c in ws[1]:
        c.font = Font(bold=True)

    teller: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    titels: dict[str, str] = {}
    for r in rijen:
        sleutel = (r["clausule"], r["norm"])
        teller[sleutel][r["classificatie"]] += 1
        if r.get("clausule_titel"):
            titels[r["clausule"]] = r["clausule_titel"]

    for (clausule, norm), t in sorted(teller.items(), key=lambda kv: _clausule_sleutel(kv[0][0])):
        ws.append(
            [
                clausule,
                norm,
                titels.get(clausule, ""),
                t.get("NC", 0),
                t.get("OFI", 0),
                t.get("positief", 0),
                sum(t.values()),
            ]
        )
    ws.freeze_panes = "A2"
    _auto_width(ws, max_breedte=50)


def _auto_width(ws: Any, max_breedte: int = 40) -> None:
    for kolom_cellen in ws.columns:
        max_lengte = 0
        letter: str | None = None
        for cel in kolom_cellen:
            if letter is None:
                letter = cel.column_letter
            waarde = str(cel.value) if cel.value is not None else ""
            eerste_regel = waarde.split("\n", 1)[0]
            if len(eerste_regel) > max_lengte:
                max_lengte = len(eerste_regel)
        if letter:
            ws.column_dimensions[letter].width = min(max_lengte + 2, max_breedte)


def lees_uit_db(db_pad: str | None = None, norm: str | None = None) -> list[dict[str, Any]]:
    """Lees bevindingen uit `audit.db` en normaliseer naar in-memory schema."""
    db_pad = db_pad or os.environ.get("AUDIT_DB_PATH", DEFAULT_DB_PATH)
    if not Path(db_pad).is_file():
        raise FileNotFoundError(f"audit.db niet gevonden: {db_pad}")

    conn = sqlite3.connect(db_pad)
    conn.row_factory = sqlite3.Row
    try:
        query = "SELECT * FROM bevindingen"
        params: tuple[Any, ...] = ()
        if norm and norm != "beide":
            query += " WHERE norm = ?"
            params = (norm,)
        rows = conn.execute(query, params).fetchall()
    finally:
        conn.close()

    return [
        {
            "id": r["id"],
            "clausule": r["clausule_id"],
            "clausule_id": r["clausule_id"],
            "clausule_titel": "",
            "norm": r["norm"],
            "classificatie": r["classificatie"],
            "beschrijving": r["beschrijving"] or "",
            "onderbouwing": r["onderbouwing"] or "",
            "document_naam": r["document_naam"] or "",
            "herkomst": r["herkomst"],
            "doc_id": r["doc_id"],
            "classified_at": r["classified_at"],
        }
        for r in rows
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="Tabulaire export van audit-bevindingen")
    parser.add_argument("--norm", choices=["9001", "27001", "beide"], default="beide")
    parser.add_argument("--scherpte", type=float, default=1.0)
    parser.add_argument("--db", default=None, help="Pad naar audit.db")
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--alleen-csv", action="store_true")
    parser.add_argument("--alleen-excel", action="store_true")
    parser.add_argument(
        "--thema-llm",
        action="store_true",
        help="Gebruik LLM voor thema-toekenning bij heuristiek-'Overig'",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    bevindingen = lees_uit_db(args.db, args.norm)
    logger.info("%d bevindingen geladen uit DB", len(bevindingen))

    llm_themas: dict[str, str] = {}
    if args.thema_llm:
        from iso_audit.classification.thema import verfijn_overig

        llm_themas = verfijn_overig(bevindingen)

    if not args.alleen_excel:
        schrijf_csv(
            bevindingen,
            args.norm,
            args.scherpte,
            args.output_dir,
            llm_themas=llm_themas,
        )
    if not args.alleen_csv:
        schrijf_excel(
            bevindingen,
            args.norm,
            args.scherpte,
            args.output_dir,
            llm_themas=llm_themas,
        )
    return 0
