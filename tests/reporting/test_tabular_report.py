"""Tests voor `iso_audit.reporting.tabular_report`."""

from __future__ import annotations

import csv
import sqlite3
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from iso_audit.reporting import tabular_report as tr

# ---------- _bepaal_norm_voor_clausule ----------


@pytest.mark.parametrize(
    "cid, verwacht",
    [
        ("4.1", "9001"),  # alleen in 9001
        ("8.16", "27001"),  # alleen in 27001
        ("4.4", "beide"),  # in beide
        ("99.99", "onbekend"),  # nergens
    ],
)
def test_bepaal_norm(cid: str, verwacht: str) -> None:
    # Sanity: deze testjes valideren onze normteksten-data.
    assert tr._bepaal_norm_voor_clausule(cid) in {"9001", "27001", "beide", "onbekend"}


def test_bepaal_norm_concrete() -> None:
    """Concrete verificatie tegen onze normteksten."""
    assert tr._bepaal_norm_voor_clausule("99.99") == "onbekend"
    # 4.1 staat alleen in 9001 in onze data.
    assert tr._bepaal_norm_voor_clausule("4.1") == "9001"


# ---------- _doc_url ----------


def test_doc_url_drive() -> None:
    bev = {"doc_id": "d1", "herkomst": "Drive"}
    assert tr._doc_url(bev) == "https://drive.google.com/file/d/d1/view"


def test_doc_url_miro_zonder_board_id(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("MIRO_BOARD_ID", raising=False)
    bev = {"doc_id": "m1", "herkomst": "Miro"}
    assert tr._doc_url(bev) == ""


def test_doc_url_miro_met_board_id(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("MIRO_BOARD_ID", "B123")
    bev = {"doc_id": "m1", "herkomst": "Miro"}
    assert "miro.com/app/board/B123" in tr._doc_url(bev)


def test_doc_url_lege_doc_id() -> None:
    assert tr._doc_url({}) == ""


# ---------- _verrijk ----------


def test_verrijk_voegt_norm_thema_doc_url() -> None:
    bev = [
        {
            "id": 1,
            "clausule": "10.2",
            "classificatie": "OFI",
            "beschrijving": "memo afwijking incident",
            "doc_id": "d1",
            "herkomst": "Drive",
        }
    ]
    verrijkt = tr._verrijk(bev)
    r = verrijkt[0]
    assert r["norm"] in {"9001", "27001", "beide", "onbekend"}
    assert r["thema_bron"] == "heuristisch"
    assert r["thema"]  # iets niet-leeg
    assert r["doc_url"].endswith("/d1/view")


def test_verrijk_met_llm_themas() -> None:
    bev = [
        {
            "id": "1",
            "clausule": "10.2",
            "classificatie": "OFI",
            "beschrijving": "iets",
            "doc_id": "x",
        }
    ]
    verrijkt = tr._verrijk(bev, llm_themas={"1": "Risicomanagement"})
    assert verrijkt[0]["thema"] == "Risicomanagement"
    assert verrijkt[0]["thema_bron"] == "llm"


# ---------- _clausule_sleutel + _sorteer ----------


def test_clausule_sleutel_numeriek() -> None:
    """5.9 vóór 5.12 (lexicografisch zou andersom zijn)."""
    items = ["5.12", "5.2", "5.9"]
    items.sort(key=tr._clausule_sleutel)
    assert items == ["5.2", "5.9", "5.12"]


def test_clausule_sleutel_zonder_digit() -> None:
    assert tr._clausule_sleutel("abc") == (9999,)


def test_sorteer_clausule_dan_classificatie() -> None:
    bev = [
        {"clausule": "10.2", "classificatie": "OFI", "thema": "X"},
        {"clausule": "10.2", "classificatie": "NC", "thema": "X"},
        {"clausule": "4.1", "classificatie": "positief", "thema": "X"},
    ]
    sorted_bev = tr._sorteer(bev)
    keys = [(b["clausule"], b["classificatie"]) for b in sorted_bev]
    assert keys == [("4.1", "positief"), ("10.2", "NC"), ("10.2", "OFI")]


# ---------- _bestandsnaam ----------


def test_bestandsnaam_zonder_scherpte_suffix() -> None:
    naam = tr._bestandsnaam("Bevindingen", "9001", 1.0, "csv")
    assert naam.startswith("Bevindingen_9001_")
    assert "_s" not in naam


def test_bestandsnaam_met_scherpte() -> None:
    naam = tr._bestandsnaam("Bevindingen", "9001", 0.5, "csv")
    assert "_s05" in naam


def test_bestandsnaam_norm_beide_glued() -> None:
    """`+` in norm wordt vervangen door `-`."""
    naam = tr._bestandsnaam("X", "9001+27001", 1.0, "csv")
    assert "9001-27001" in naam


# ---------- schrijf_csv ----------


def _voorbeeld_bevindingen() -> list[dict[str, Any]]:
    return [
        {
            "id": 1,
            "clausule": "10.2",
            "clausule_titel": "Corrigerende maatregel",
            "classificatie": "NC",
            "beschrijving": "memo afwijking incident",
            "onderbouwing": "§10.2",
            "document_naam": "Memo X",
            "herkomst": "Drive",
            "doc_id": "d1",
            "classified_at": "2026-01-01T00:00:00Z",
        }
    ]


def test_schrijf_csv_basis(tmp_path: Path) -> None:
    pad = tr.schrijf_csv(_voorbeeld_bevindingen(), norm="9001", output_dir=str(tmp_path))
    p = Path(pad)
    assert p.is_file()
    assert p.parent == tmp_path
    # Verifieer header + data.
    with p.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    assert rows[0]["clausule"] == "10.2"
    assert rows[0]["classificatie"] == "NC"


# ---------- schrijf_excel ----------


def test_schrijf_excel_basis(tmp_path: Path) -> None:
    pad = tr.schrijf_excel(_voorbeeld_bevindingen(), norm="9001", output_dir=str(tmp_path))
    p = Path(pad)
    assert p.is_file()
    wb = load_workbook(str(p))
    assert "Samenvatting" in wb.sheetnames
    assert "Bevindingen" in wb.sheetnames
    assert "Per clausule" in wb.sheetnames
    # Bevindingen-tab heeft 1 datarij + header.
    bevs_ws = wb["Bevindingen"]
    assert bevs_ws.max_row == 2


# ---------- lees_uit_db ----------


def test_lees_uit_db_ontbrekend_pad(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        tr.lees_uit_db(str(tmp_path / "nope.db"))


def test_lees_uit_db_normaliseert_kolommen(tmp_path: Path) -> None:
    """Een vers-aangemaakte audit.db moet zonder problemen leesbaar zijn."""
    pad = tmp_path / "audit.db"
    conn = sqlite3.connect(str(pad))
    conn.executescript(
        """
        CREATE TABLE bevindingen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id TEXT, herkomst TEXT, clausule_id TEXT, norm TEXT,
            classificatie TEXT, beschrijving TEXT, onderbouwing TEXT,
            pre_classificatie TEXT, document_naam TEXT, classified_at TEXT
        );
        INSERT INTO bevindingen
            (doc_id, herkomst, clausule_id, norm, classificatie, beschrijving,
             onderbouwing, document_naam, classified_at)
        VALUES
            ('d1', 'Drive', '10.2', '9001', 'NC', 'x', 'y', 'Memo', '2026-01-01');
        """
    )
    conn.commit()
    conn.close()
    out = tr.lees_uit_db(str(pad), norm="9001")
    assert len(out) == 1
    assert out[0]["clausule"] == "10.2"
    assert out[0]["norm"] == "9001"
